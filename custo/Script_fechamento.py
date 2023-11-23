from openpyxl import load_workbook, Workbook
import os

#Script Para salvar o relatório tratado na planilha de Fechamento de custo

# Colunas do novo cabeçalho
novo_cabecalho = ['Material', 'Texto breve material', 'Estoque total', 'UMB', 'Val.total', 'Tipo', 'Classe', 'Custo Total', 'Total MP', 'Total IMP']

# Caminho para os arquivos Excel
caminho_arquivo_origem = './novo.xlsx'
#caminho_arquivo_destino = 'C:/Users/lcrodrigues/Documents/Indicadores_Almox/Indicadores_considerados/Fechamento.xlsx'
caminho_arquivo_destino = ':\Users\lcrodrigues\Documents\custo\Fechamento.xls'


# Carrega os arquivos Excel
arquivo_origem = load_workbook(filename=caminho_arquivo_origem)
arquivo_destino = load_workbook(filename=caminho_arquivo_destino)

# Seleciona a aba de origem e destino
aba_origem = arquivo_origem.active
aba_destino = arquivo_destino['Custo']

# Cria o novo cabeçalho na aba de destino (se não existir)
if aba_destino.max_row == 0:
    aba_destino.append(novo_cabecalho)

# Apaga dados existentes na aba de destino, se houver (exceto o cabeçalho)
for _ in range(2, aba_destino.max_row + 1):
    aba_destino.delete_rows(2)

# Copia os dados da aba de origem para a aba de destino
for linha in aba_origem.iter_rows(min_row=0, values_only=True):  # Começando da primeira linha após o cabeçalho
    aba_destino.append(linha)

# Salva o arquivo de destino
arquivo_destino.save(caminho_arquivo_destino)
