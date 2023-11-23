import pandas as pd 
from openpyxl import load_workbook

def carregar_dados(arquivo):
    excel_file = pd.ExcelFile(arquivo)
    df = pd.read_excel(excel_file, header=6)
    return df

def limpar_dataframe(df):
    df = df.drop(index=range(5))
    df = df.dropna(axis=1, how='all')
    colunas_para_remover = ['Tp.aval.', 'Documento SD', 'Prç.padrão', 'PrçIntPer.', 'Tipo de material', 'Unnamed: 1', 'Moeda']
    df = df.drop(columns=colunas_para_remover, errors='ignore')
    df.columns = df.columns.str.strip()
    ordem_colunas = ['Material', 'Texto breve material','Estoque total', 'UMB', 'Val.total']
    df = df[ordem_colunas]
    indice_linha_vazia = df['Material'][df['Material'].isnull() & df['Material'].shift(-1).isnull()].index.min()
    if pd.notnull(indice_linha_vazia):
        df = df.loc[:indice_linha_vazia]
    return df

def merge_dataframes(df, df_classe, df_remover):
    df = pd.merge(df, df_classe[['Material', 'Tipo', 'Classe']], on='Material', how='left')
    merged = pd.merge(df, df_remover[['Material']], on='Material', how='left', indicator=True)
    df_resultado = merged[merged['_merge'] == 'left_only'].drop('_merge', axis=1)
    return df_resultado

def calcular_totais(df_resultado):
    df_resultado['Tipo'] = df_resultado['Tipo'].str.capitalize()
    total = df_resultado.groupby('Tipo')['Val.total'].sum()
    df_resultado.at[0, 'Custo Total'] = total.sum()
    df_resultado.at[0, 'Total MP'] = total.get('Materia prima', 0)
    df_resultado.at[0, 'Total IMP'] = total.get('Improdutivo', 0)
    return df_resultado

def salvar_resultados(df, caminho):
    df.to_excel(caminho, index=False)

def atualizar_arquivo_excel(novo_cabecalho, caminho_arquivo_origem, caminho_arquivo_destino):
    arquivo_origem = load_workbook(filename=caminho_arquivo_origem)
    arquivo_destino = load_workbook(filename=caminho_arquivo_destino)

    aba_origem = arquivo_origem.active
    aba_destino = arquivo_destino['Custo']

    if aba_destino.max_row == 0:
        aba_destino.append(novo_cabecalho)

    for _ in range(2, aba_destino.max_row + 1):
        aba_destino.delete_rows(2)

    for linha in aba_origem.iter_rows(min_row=0, values_only=True):
        aba_destino.append(linha)

    arquivo_destino.save(caminho_arquivo_destino)

def main():
    arquivo_dados = './custo/custo.xlsx'
    arquivo_classe = './custo/classe.xlsx'
    arquivo_remover = './custo/remover.xlsx'
    arquivo_destino = './custo/novo.xlsx'
    caminho_arquivo_origem = './custo/novo.xlsx'
    caminho_arquivo_destino = './Indicadores_Almox/Indicadores_considerados/Fechamento.xlsx'
    
    df = carregar_dados(arquivo_dados)
    df = limpar_dataframe(df)
    
    df_classe = carregar_dados(arquivo_classe)
    df_resultado = merge_dataframes(df, df_classe, carregar_dados(arquivo_remover))
    
    df_resultado = calcular_totais(df_resultado)
    salvar_resultados(df_resultado, arquivo_destino)
    
    novo_cabecalho = ['Material', 'Texto breve material', 'Estoque total', 'UMB', 'Val.total', 'Tipo', 'Classe', 'Custo Total', 'Total MP', 'Total IMP']
    atualizar_arquivo_excel(novo_cabecalho, caminho_arquivo_origem, caminho_arquivo_destino)

if __name__ == "__main__":
    main()
